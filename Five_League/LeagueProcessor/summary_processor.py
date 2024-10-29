import numpy as np
import pandas as pd
import re
# from .normal_channel_result_processor import process_normal_result
from .data_cleaner import DataCleaner
from .day_span_processor import DaySpanProcessor
from .file_manager import FileManager

import sys
import os

parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
from .other_func import *
from .support_files import *
import emoji
import time
import warnings
import shutil
warnings.filterwarnings('ignore')
from openpyxl import Workbook

class SummaryProcessor:
    def __init__(self, shared_variables):
        self.update_date = shared_variables["Period"]
        self.overall_date = shared_variables["Overall"]
    
    def get_updated_df(self, df, period, league=""):
        # 找到 title_row 的位置
        title_row = df.apply(lambda x: any(x.str.contains("date", case=False, na=False)) and any(x.str.contains("start", case=False, na=False)), axis=1).idxmax()
        
        # 定义 target_row
        target_row = title_row - 1
        
        # 转置 target_row 行并填充数据
        tp_target_df = pd.DataFrame(df.iloc[target_row].values).T
        
        tp_target_df_filled = tp_target_df.ffill(axis=1)
        
        df.iloc[target_row] = tp_target_df_filled.values.flatten()
        
        # 定义 tv_start_col
        tv_start_col = np.where(df.iloc[target_row].str.contains("Targets", na=False))[0][0] + 1
        
        # 构建 ColNames
        ColNames = list(df.iloc[title_row, :tv_start_col]) + \
                   [f"{df.iloc[target_row, i]} {df.iloc[title_row, i]}" for i in range(tv_start_col, df.shape[1])]
        
        # 设置列名
        df.columns = ColNames
        
        # 删除 title_row 及之前的行
        df = df.iloc[title_row+1:].copy()
        
        df = df[
            (df['Date'] >= pd.to_datetime(period[0])) & 
            (df['Date'] <= pd.to_datetime(period[1]))
        ]
        
        # 选择前 21 列并添加 League 列
        df = df.iloc[:, :22].copy().reset_index(drop=True)
        df['League'] = league
            
        return df

    def get_df_all(self, df_epl, df_ger, df_fra, df_esp, df_ita):
        # 合并数据框
        df_new = pd.concat([df_epl, df_ger, df_fra, df_esp, df_ita], ignore_index=True)
        
        # 转换指定列的数据类型
        numeric_columns = ['Dur.(sec)', 'All 4 years + TVR(000)', 'All 4 years + Rch(000)']
        df_new[numeric_columns] = df_new[numeric_columns].apply(pd.to_numeric, errors='coerce')
        
        # 设置 League 列的类别顺序
        league_order = pd.CategoricalDtype(["EPL", "La Liga", "Bundesliga", "Ligue 1", "Serie A"], ordered=True)
        df_new['League'] = df_new['League'].astype(league_order)
        
        # 按 League 和 Channel 排序
        df_new = df_new.sort_values(by=['League', 'Channel']).reset_index(drop=True)
        
        return df_new

    def get_summary(self, df_new, league_lists=["EPL", "La Liga", "Bundesliga", "Ligue 1", "Serie A", "Bundesliga only CCTVs"]):
        # 按 League 分组并计算汇总数据
        df_summ = df_new.groupby('League').agg(
            nn=('League', 'size'),
            total_time=('Dur.(sec)', lambda x: x.sum() / (24 * 3600)),  # 将总时间转换为天
            cul_tvr=('All 4 years + TVR(000)', 'sum'),
            cul_rch=('All 4 years + Rch(000)', 'sum')
        ).reset_index()
    
        # 处理 Bundesliga 且 Channel 包含 CCTV 的特殊情况
        df_special = df_new[(df_new['League'] == "Bundesliga") & (df_new['Channel'].str.contains("CCTV"))].copy()
        df_special['League'] = "Bundesliga only CCTVs"
        df_special = df_special.groupby('League').agg(
            nn=('League', 'size'),
            total_time=('Dur.(sec)', lambda x: x.sum() / (24 * 3600)),
            cul_tvr=('All 4 years + TVR(000)', 'sum'),
            cul_rch=('All 4 years + Rch(000)', 'sum')
        ).reset_index()
    
        # 合并一般情况和特殊情况的汇总
        df_summ = pd.concat([df_summ, df_special], ignore_index=True)
    
        # 检查 league_lists 中是否有不存在的 League，如果有则添加空值行
        for ll in league_lists:
            if ll not in df_summ['League'].values:
                df_summ = pd.concat([
                    df_summ,
                    pd.DataFrame({
                        'League': [ll],
                        'nn': [pd.NA],
                        'total_time': [pd.NA],
                        'cul_tvr': [pd.NA],
                        'cul_rch': [pd.NA]
                    })
                ], ignore_index=True)
    
        # 按 League 排序
        # df_summ = df_summ.sort_values(by='League').reset_index(drop=True)
        
        return df_summ
    
    def get_fra_report_summ(self, df_new):
        # 设置 League 列的类别顺序
        league_order = pd.CategoricalDtype(["EPL", "Ligue 1", "Bundesliga", "La Liga", "Serie A"], ordered=True)
        df_new['League'] = df_new['League'].astype(league_order)
        
        # df_fra_report 的生成
        df_fra_report = df_new.copy()
        df_fra_report['rn'] = df_fra_report.groupby(['Round', 'League', 'Channel', 'Teams Detail(Eng)']).cumcount() + 1
        df_fra_report = df_fra_report[((df_fra_report['rn'] == 1) & (df_fra_report['Air type'] != "Live")) | (df_fra_report['Air type'] == "Live")]
        
        df_fra_report['Air type'] = df_fra_report['Air type'].apply(lambda x: "Live" if x == "Live" else "Delay")
        df_fra_report['Air type'] = pd.Categorical(df_fra_report['Air type'], categories=["Live", "Delay"], ordered=True)
        
        df_fra_report = df_fra_report.groupby(['League', 'Air type', 'Channel']).agg(
            nn=('Air type', 'size'),
            total_time=('Dur.(sec)', lambda x: x.sum() / (24 * 3600)),
            cul_tvr=('All 4 years + TVR(000)', 'sum'),
            cul_rch=('All 4 years + Rch(000)', 'sum')
        ).reset_index()
        
        df_fra_report = df_fra_report.sort_values(by=['League', 'Air type', 'cul_tvr'], ascending=[True, True, False])
        
        df_fra_report = df_fra_report[df_fra_report['total_time'].notna()].copy().reset_index()
        
        # 去重处理
        df_fra_report['Air type'] = df_fra_report['Air type'].mask(df_fra_report.duplicated(['League', 'Air type']))
        df_fra_report['League'] = df_fra_report['League'].mask(df_fra_report.duplicated('League'))
        
        df_new['League'] = df_new['League'].astype(str)
        
        # df_other_top_live 的生成
        df_other_top_live = df_new[(df_new['Air type'] == "Live") & (df_new['League'] != "Ligue 1")].copy()
        df_other_top_live = df_other_top_live.groupby(['League', 'Round', 'Teams Detail(Eng)']).agg(
            nn=('Air type', 'size'),
            cul_tvr=('All 4 years + TVR(000)', 'sum')
        ).reset_index()
        
        df_other_top_live = df_other_top_live.loc[df_other_top_live.groupby(['League', 'Round'])['cul_tvr'].idxmax()]
        
        df_other_top_live = df_other_top_live.merge(
            df_new[(df_new['Air type'] == "Live") & (df_new['League'] != "Ligue 1")],
            on=['League', 'Round', 'Teams Detail(Eng)'],
            how='left'
        )
        
        df_other_top_live = df_other_top_live[['League', 'Round', 'Teams Detail(Eng)', 'Date', 'Day', 'Live Timeslot', 'nn', 'cul_tvr']].drop_duplicates()
        
        # df_other_top_first 的生成
        df_other_top_first = df_new[df_new['League'] != "Ligue 1"].copy()
        
        df_other_top_first['rn'] = df_other_top_first.groupby(['League', 'Round', 'Teams Detail(Eng)', 'Channel']).cumcount() + 1
        
        df_other_top_first = df_other_top_first.groupby(['League', 'Round', 'Teams Detail(Eng)', 'Channel']).apply(lambda x: x.loc[x['rn'].idxmin()], include_groups=False).reset_index()
        
        df_other_top_first = df_other_top_first.groupby(['League', 'Round', 'Teams Detail(Eng)']).agg(
            nn=('Channel', 'size'),
            cul_tvr=('All 4 years + TVR(000)', 'sum')
        ).reset_index()
        
        df_other_top_first = df_other_top_first.loc[df_other_top_first.groupby(['League', 'Round'])['cul_tvr'].idxmax()]
        
        df_other_top_first.loc[df_other_top_first.groupby(['League'])['cul_tvr'].idxmax()]
        
        return [df_fra_report, df_other_top_live, df_other_top_first]
  
    def process_summary_result(self):
        update_date = self.update_date
        overall_date = self.overall_date
        
        file_name = update_date[1].strftime("%Y%m%d")
        
        league_lists = ["EPL", "La Liga", "Bundesliga", "Ligue 1", "Serie A", "Bundesliga only CCTVs"]
        
        # 读取报告文件
        print("Reading report files......")
        base_path = f"D:/wangxiaoyang/Regular_Work/Produce_Report/Reports/Five_League/{file_name}/"
        files = os.listdir(base_path)
        
        file_epl = next((f for f in files if f.startswith("2024-25 English Premier League - China")), None)
        df_epl_org = pd.read_excel(os.path.join(base_path, file_epl), sheet_name="TV Audience Report")
        
        file_ger = next((f for f in files if f.startswith("Bundesliga 2024-25")), None)
        df_ger_org = pd.read_excel(os.path.join(base_path, file_ger), sheet_name="Bundesliga TV Audience Report")
        
        file_fra = next((f for f in files if f.startswith("2024-25 French Football Ligue 1")), None)
        df_fra_org = pd.read_excel(os.path.join(base_path, file_fra), sheet_name="TAM Report")
        
        file_esp = next((f for f in files if f.startswith("2024-25 La Liga")), None)
        df_esp_org = pd.read_excel(os.path.join(base_path, file_esp), sheet_name="TV Audience Report")
        
        # 更新汇总
        print("Generating summary for update data......")
        df_epl = self.get_updated_df(df_epl_org, update_date, "EPL")
        df_ger = self.get_updated_df(df_ger_org, update_date, "Bundesliga")
        df_fra = self.get_updated_df(df_fra_org, update_date, "Ligue 1")
        df_esp = self.get_updated_df(df_esp_org, update_date, "La Liga")
        
        df_summ_new = self.get_summary(self.get_df_all(df_epl, df_ger, df_fra, df_esp, None))
        
        # 总体汇总
        print("Generating summary for all data......")
        df_epl = self.get_updated_df(df_epl_org, overall_date, "EPL")
        df_ger = self.get_updated_df(df_ger_org, overall_date, "Bundesliga")
        df_fra = self.get_updated_df(df_fra_org, overall_date, "Ligue 1")
        df_esp = self.get_updated_df(df_esp_org, overall_date, "La Liga")
        
        df_summ = self.get_summary(self.get_df_all(df_epl, df_ger, df_fra, df_esp, None))
        
        # 法甲报告总结
        print("Generating summary for French Report......")
        df_fra_summ = self.get_fra_report_summ(self.get_df_all(df_epl, df_ger, df_fra, df_esp, None))
        
        # 导出结果
        wb = Workbook()
        wb.create_sheet("update")
        wb["update"].append(df_summ_new.columns.tolist())
        for row in df_summ_new.values:
            wb["update"].append(row.tolist())
        
        wb.create_sheet("overall")
        wb["overall"].append(df_summ.columns.tolist())
        for row in df_summ.values:
            wb["overall"].append(row.tolist())
        
        wb.create_sheet("Sheet1")
        wb["Sheet1"].append(df_fra_summ[0].columns.tolist())
        for row in df_fra_summ[0].values:
            wb["Sheet1"].append(row.tolist())
        
        wb.create_sheet("Sheet2")
        wb["Sheet2"].append(df_fra_summ[1].columns.tolist())
        for row in df_fra_summ[1].values:
            wb["Sheet2"].append(row.tolist())
        
        wb.create_sheet("Sheet3")
        wb["Sheet3"].append(df_fra_summ[2].columns.tolist())
        for row in df_fra_summ[2].values:
            wb["Sheet3"].append(row.tolist())
        
        wb.save(f"{parent_dir}/Result/result_summary.xlsx")
        
        # 打开文件
        os.startfile(f"D:/wangxiaoyang/Regular_Work/Produce_Report/Reports/Five_League/{file_name}/{file_ger}")
        os.startfile(f"D:/wangxiaoyang/Regular_Work/Produce_Report/Reports/Five_League/{file_name}/{file_fra}")
        os.startfile(f"{parent_dir}/Result/result_summary.xlsx")
        
        # 法甲 UniRch 数据生成
        print("Generating Infosys Program Sheet txt file for French UniRch......")
        df_fra_unirch = df_fra.assign(
                    Start=df_fra['Start'].apply(lambda x: convert_time( get_seconds(x.strftime("%H:%M:%S"), False) )),
                    End=df_fra['End'].apply(lambda x: convert_time( get_seconds(x.strftime("%H:%M:%S"), False) ))
                ).merge(
                    channel_mapping,
                    left_on='Channel',
                    right_on='channel',
                    how='left'
                ).assign(
                    ProgramID=lambda x: x['Channel'] + " " + x['Date'].astype(str) + " " + x['Start']
                )
        
        # 拆分跨天数据
        df_fra_unirch1 = df_fra_unirch[df_fra_unirch['Start'] <= df_fra_unirch['End']]
        df_fra_unirch2 = df_fra_unirch[df_fra_unirch['Start'] > df_fra_unirch['End']]
        
        # 创建跨天数据帧
        df_fra_ds = pd.concat([
            df_fra_unirch2.assign(End="25:59:59"),
            df_fra_unirch2.assign(Start="02:00:00")
        ])
        
        df_fra_unirch = pd.concat([df_fra_unirch1, df_fra_ds])
        
        # 生成目标格式并写入文件
        res = df_fra_unirch.apply(lambda row: f"{row['Date'].strftime('%Y%m%d')} {str(row['code']).zfill(4)} {row['Start'].replace(':', '')} {row['End'].replace(':', '')} {row['ProgramID']}", axis=1)
        
        with open(f"{parent_dir}/infosys_txt/Ligue 1_unirch_ps.txt", "w", encoding="GBK") as f:
            for line in res:
                f.write(line + "\n")
        
        print("Done.")
