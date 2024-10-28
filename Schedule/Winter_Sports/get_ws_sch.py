import requests
from bs4 import BeautifulSoup
import pandas as pd
import datetime
import pytz
from tqdm import tqdm
import os
from openpyxl import load_workbook
import argparse

def convert_tz(datetime_str, tz_str = 'Asia/Shanghai', tz_org = None):
    if datetime_str:
        datetime_obj = pd.to_datetime(datetime_str)
        tz = pytz.timezone(tz_str)

        if tz_org:
            datetime_obj = datetime_obj.tz_localize(tz_org)
        datetime_dt = datetime_obj.astimezone(tz)
    else:
        datetime_dt = None
    return datetime_dt
    # datetime_dt.strftime("%Y-%m-%d %H:%M:%S")


def convert_to_excel_time(time_str):
    time_parts = [int(part) for part in time_str.split(":")]
    total_seconds = time_parts[0] * 3600 + time_parts[1] * 60 + (time_parts[2] if len(time_parts) == 3 else 0)
    return total_seconds / (24*3600)

comp_mapping = pd.read_excel("./ws_mapping.xlsx")
tz_mapping = pd.read_csv("D:/wangxiaoyang/Regular_Work/support_files/time_zone_population.csv")

parser = argparse.ArgumentParser(description='Get winter sports schedule.')
parser.add_argument('-s', '--season', type=str, required=True, help='Season')
args = parser.parse_args()

season = args.season

# Fetch the page
def get_sch_skiing(comp, cate, season):
    url = f'https://www.fis-ski.com/DB/general/calendar-results.html?eventselection=&place=&sectorcode={comp}&seasoncode={season}&categorycode={cate}&disciplinecode=&gendercode=&racedate=&racecodex=&nationcode=&seasonmonth=X-{season}&saveselection=-1&seasonselection='
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Find all event rows
    rows = soup.find_all('div', class_='table-row')
    
    df = pd.DataFrame()
    # Extract event data
    # i = 1
    for row in tqdm(rows, desc=f"Crawling {comp}-{cate}-{season}", unit="competition"):
        # print(f"{i} / {len(rows)}")
        # i+=1
        link = row.find('a', href=True)
        if link:
            href = link['href']
        else:
            print("error")
            continue
        response_new = requests.get(href)
        soup_new = BeautifulSoup(response_new.content, 'html.parser')
    
        venue = soup_new.find('title').get_text(strip=True)
        # Find all event rows
        rows_new = soup_new.find_all('div', class_='table-row')
    
        for rr in rows_new:
    
            dd = rr.find('div', class_='timezone-date')
            if dd:
                date = dd['data-date']
                time = dd['data-time']
                ddtt = dd['data-iso-date']
                discipline = rr.find('div', class_='clip').get_text(strip=True)
                gender = rr.find('div', class_='gender__inner').get_text(strip=True)
                cancelled = rr.find('div', class_='g-row').get_text(strip=True)
            df_tp = pd.DataFrame(
                {
                    'Event': f"{comp}-{cate}",
                    'Gender': gender,
                    'Date': date,
                    'Venue': venue,
                    'Discipline': discipline,
                    'Start': time,
                    'ISO': ddtt,
                    'Status': cancelled
                }, index = [0]
            )
            df = pd.concat([df, df_tp])
    return df

def get_sch_curling(comp):
    global tz_org
    url = f'https://livescores.worldcurling.org/{comp}/aspnet/summary?EventID=1'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    df = pd.DataFrame()
    
    temp = soup.find('span', id="LabelPlace").get_text(strip=True).split(', ')
    country = temp[len(temp)-1]
    tz_org = tz_mapping.loc[tz_mapping['CSM.Territory'] == country, 'IANA.TZ.identifier'].iloc[0]
    
    rows = soup.find_all('tr', valign='top')
    for row in tqdm(rows, desc=f"Crawling {comp}", unit="competition"):
        cells = row.find_all('td')
        stage = cells[0].get_text(strip=True).replace('\xa0', '')
        date_time = cells[1].get_text(strip=True)
        nums = int((len(cells)-1) / 3)
        data_tp_list = []
        for i in range(1, nums+1):
            # print(i)
            teams = cells[i*3-1].get_text().replace('\xa0', '').split(' ')
            team_1, team_2 = teams[0], teams[1]
            data_tp = {
                'Competition': comp,
                'Stage': stage,
                'time_org': date_time,
                'Team_1': team_1,
                'Team_2': team_2
            }
            data_tp_list.append(data_tp)
        df_tp = pd.DataFrame(data_tp_list)
        df = pd.concat([df, df_tp])
    return df

def deal_sch_skiing(df):
    df1 = df.drop_duplicates().reset_index(drop=True)
    df1['Date'] = df1.apply(
        lambda row: convert_tz(row['ISO']).strftime("%Y-%m-%d") if convert_tz(row['ISO']) else row['Date'],
        axis=1
    )

    df1['Start'] = df1.apply(
        lambda row: convert_to_excel_time(convert_tz(row['ISO']).strftime("%H:%M:%S")) if convert_tz(row['ISO']) else row['Start'],
        axis=1
    )
    
    df1['Status'] = df1.apply(
        lambda row: 'cancelled' if 'Cancel' in row['Status'] else None,
        axis=1
    )
    
    df1['Gender'] = df1['Gender'].apply(
        lambda x: 'Men' if x=='M' else 'Women'
    )
    
    df1['Venue'] = df1['Venue'].str.split(' - ').str[-2]
    
    # Extracting the nation code, which is in parentheses, and creating a new 'Nation' column
    df1['Nation'] = df1['Venue'].str.extract(r'\((\w{3})\)')
    
    df1 = df1.drop(columns=['ISO'])
    
    df2 = df1.merge(comp_mapping, how='left', left_on='Event', right_on='code')
    
    df2['Event'] = df2['name']
    df2 = df2.drop(columns=['name', 'code'])
    return df2

def deal_sch_curling(df):
    tz = pytz.timezone(tz_org)
    df['Time'] = df['time_org'].apply(
        lambda x: convert_tz(datetime.datetime.strptime(f"2024 {x}", "%Y %a %d %b%H:%M").strftime("%Y-%m-%d %H:%M:%S"), tz_org = tz_org)
    )
    df1 = df[df['Team_1'] != ''].reset_index(drop=True)

    df1['Date'] = df1['Time'].dt.strftime("%Y-%m-%d")
    
    df1['Start'] = df1.apply(
        lambda row: convert_to_excel_time(row['Time'].strftime("%H:%M:%S")),
        axis=1
    )
    
    df2 = df1[['Competition', 'Stage', 'Date', 'Team_1', 'Team_2', 'Start']]
    return df2

df_wc = pd.DataFrame()
for cc in tqdm(['AL', 'CC', 'NK', 'JP', 'SB', 'FS'], desc=f"Crawling Skiing World Cup Schedule", unit="discipline"):
    df_tp = get_sch_skiing(cc, 'WC', season)
    if len(df_tp) == 0:
        continue
    df_tp = deal_sch_skiing(df_tp)
    df_wc = pd.concat([df_wc, df_tp])

df_wsc = pd.DataFrame()
for cc in tqdm(['AL', 'NK'], desc=f"Crawling Skiing World Championships Schedule", unit="discipline"):
    df_tp = get_sch_skiing(cc, 'WSC', season)
    if len(df_tp) == 0:
        continue
    df_tp = deal_sch_skiing(df_tp)
    df_wsc = pd.concat([df_wsc, df_tp])

df_res_skiing = pd.concat([df_wc, df_wsc]).reset_index(drop=True)
df_res_skiing['Date'] = pd.to_datetime(df_res_skiing['Date'])
# df_res.to_excel(f"./schedule_{season}.xlsx", index=False)

## Curling ====
df_res = pd.DataFrame()
for cc in tqdm(['wmxcc', 'pccc', 'ecc'], desc=f"Crawling Curling Schedule", unit="discipline"):
    df_tp = get_sch_curling(cc)
    if len(df_tp) == 0:
        continue
    df_tp = deal_sch_curling(df_tp)
    df_res = pd.concat([df_res, df_tp])
tqdm.write("Done crawling Curling schedule.")
df_res = df_res.merge(comp_mapping, how='left', left_on='Competition', right_on='code')
df_res['Competition'] = df_res['name']
df_res = df_res.drop(columns=['name', 'code'])
df_res['Date'] = pd.to_datetime(df_res['Date'])

file_path = f"./schedule-{season}.xlsx"
if not os.path.exists(file_path):
    with pd.ExcelWriter(f"./schedule-{season}.xlsx", mode='w', engine='openpyxl') as writer:
        df_res_skiing.to_excel(writer, sheet_name="Skiing", index=False)
else:
    with pd.ExcelWriter(f"./schedule-{season}.xlsx", mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_res_skiing.to_excel(writer, sheet_name="Skiing", index=False)


if not os.path.exists(file_path):
    with pd.ExcelWriter(f"./schedule-{season}.xlsx", mode='w', engine='openpyxl') as writer:
        df_res.to_excel(writer, sheet_name="Curling", index=False)
else:
    with pd.ExcelWriter(f"./schedule-{season}.xlsx", mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_res.to_excel(writer, sheet_name="Curling", index=False) 

for comp in ["Skiing", "Curling"]:
  # Now apply the date format to the 'Date' column (after writing)
  wb = load_workbook(file_path)
  ws = wb[comp]  # Assuming you're working with the "Curling" sheet
  
  # Apply the date format to the 'Date' column (e.g., column B)
  for cell in ws['C']:  # Assuming 'Date' is in column B
      cell.number_format="YYYY/MM/DD"
  for cell in ws['F']:
      cell.number_format="HH:MM:SS"
  # Save the workbook with formatting changes
  wb.save(file_path)








# U型场地|北欧两项|国际雪联|冰壶|滑雪
# 自由式|单板|冰壶|跳台|高山|U型|北欧两项|自由身|越野
